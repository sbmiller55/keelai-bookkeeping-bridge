"""
Accrued Expenses Schedule endpoints.

Routes:
  GET    /clients/{id}/accruals                       — list accrued expenses + summary
  POST   /clients/{id}/accruals                       — create manually
  PATCH  /clients/{id}/accruals/{ae_id}               — update status / date / amount
  DELETE /clients/{id}/accruals/{ae_id}               — delete
  POST   /clients/{id}/accruals/analyze               — AI scan pending payments
  GET    /clients/{id}/accruals/standing-rules        — list standing rules
  POST   /clients/{id}/accruals/standing-rules        — create rule
  PATCH  /clients/{id}/accruals/standing-rules/{rid}  — update rule
  DELETE /clients/{id}/accruals/standing-rules/{rid}  — delete rule
  POST   /clients/{id}/accruals/standing-rules/generate — generate this month's JEs
"""

import calendar
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models import (
    AccruedExpense,
    AccruedExpenseStatus,
    BillingType,
    Client,
    FixedAsset,
    JournalEntry,
    RevenueContract,
    RevenueScheduleEntry,
    RevenueStream,
    StandingAccrualRule,
    Transaction,
    TransactionStatus,
    User,
    next_je_number,
)
from schemas import (
    AccruedExpenseCreate,
    AccruedExpenseRead,
    AccruedExpenseUpdate,
    AccrualSummary,
    StandingAccrualRuleCreate,
    StandingAccrualRuleRead,
    StandingAccrualRuleUpdate,
)

router = APIRouter(prefix="/clients/{client_id}/accruals", tags=["accruals"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_client(client_id: int, user: User, db: Session) -> Client:
    client = db.query(Client).filter(Client.id == client_id, Client.user_id == user.id).first()
    if not client:
        raise HTTPException(404, "Client not found")
    return client


def _current_month() -> str:
    return datetime.utcnow().strftime("%Y-%m")


def _derive_status(ae: AccruedExpense, je: Optional[JournalEntry], this_month: str) -> str:
    """
    Map an AccruedExpense to a user-facing status label.

      - 'cleared'    — cash payment matched (ae.status == cleared)
      - 'recognized' — accrual JE approved/exported to QBO
      - 'pending'    — JE in Review Queue, not yet approved
      - 'upcoming'   — future-month entry, no JE yet
      - 'overdue'    — past-month entry, no JE yet
    """
    if ae.status == AccruedExpenseStatus.cleared:
        return "cleared"
    if ae.accrual_je_id and je is not None:
        if je.approved_at or je.exported_at:
            return "recognized"
        return "pending"
    # No JE yet — compare period to current month
    if ae.service_period > this_month:
        return "upcoming"
    if ae.service_period < this_month:
        return "overdue"
    return "upcoming"  # current month, not yet released → upcoming


def _create_synthetic_transaction(
    client_id: int,
    vendor: str,
    amount: float,
    tx_date: datetime,
    db: Session,
    status: TransactionStatus = TransactionStatus.approved,
) -> Transaction:
    """Create a synthetic transaction to anchor an accrual JE. Callers can
    override status — standing-rule generation uses `pending` for any month
    whose end date has already arrived so the user can review the JE."""
    tx = Transaction(
        client_id=client_id,
        date=tx_date,
        description=f"Accrual: {vendor}",
        amount=-abs(amount),
        status=status,
        source="accrual",
    )
    db.add(tx)
    db.flush()
    return tx


def _months_in_range(start_ym: str, end_ym: str) -> list[str]:
    """Inclusive list of YYYY-MM strings from start to end (or empty if out of order)."""
    try:
        s = datetime.strptime(start_ym, "%Y-%m")
        e = datetime.strptime(end_ym, "%Y-%m")
    except ValueError:
        return []
    out: list[str] = []
    cur = s
    while cur <= e:
        out.append(cur.strftime("%Y-%m"))
        if cur.month == 12:
            cur = datetime(cur.year + 1, 1, 1)
        else:
            cur = datetime(cur.year, cur.month + 1, 1)
    return out


def _next_month_str(ym: str) -> str:
    dt = datetime.strptime(ym, "%Y-%m")
    if dt.month == 12:
        return f"{dt.year + 1}-01"
    return f"{dt.year}-{dt.month + 1:02d}"


def _create_accrual_je(tx_id: int, expense_account: str, accrued_account: str,
                        amount: float, je_date: datetime, vendor: str,
                        service_period: str, confidence: float, reasoning: str,
                        db: Session) -> JournalEntry:
    je = JournalEntry(
        transaction_id=tx_id,
        je_number=next_je_number(db),
        debit_account=expense_account,
        credit_account=accrued_account,
        amount=amount,
        je_date=je_date,
        memo=f"Accrual: {vendor} ({service_period})"[:80],
        ai_confidence=confidence,
        ai_reasoning=reasoning,
    )
    db.add(je)
    db.flush()
    return je


def _create_payment_je(tx_id: int, accrued_account: str, bank_account: str,
                        amount: float, pay_date: datetime, vendor: str,
                        db: Session) -> JournalEntry:
    je = JournalEntry(
        transaction_id=tx_id,
        je_number=next_je_number(db),
        debit_account=accrued_account,
        credit_account=bank_account,
        amount=amount,
        je_date=pay_date,
        memo=f"Payment: {vendor}"[:80],
        ai_confidence=1.0,
        ai_reasoning="Payment JE clearing accrued liability.",
    )
    db.add(je)
    db.flush()
    return je


# ── List & summary ────────────────────────────────────────────────────────────

@router.get("")
def list_accruals(
    client_id: int,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)
    q = db.query(AccruedExpense).filter(AccruedExpense.client_id == client_id)
    if status:
        q = q.filter(AccruedExpense.status == status)
    accruals = q.order_by(AccruedExpense.service_period.asc(), AccruedExpense.created_at.asc()).all()

    # Fetch all the linked JEs in one query for derived-status calculation
    je_ids = [a.accrual_je_id for a in accruals if a.accrual_je_id]
    je_map: dict[int, JournalEntry] = {}
    if je_ids:
        for je in db.query(JournalEntry).filter(JournalEntry.id.in_(je_ids)).all():
            je_map[je.id] = je

    this_month = _current_month()

    rows = []
    upcoming_this_month_amount = 0.0
    outstanding_accrual_amount = 0.0
    prepaid_balance = 0.0
    for a in accruals:
        je = je_map.get(a.accrual_je_id) if a.accrual_je_id else None
        derived = _derive_status(a, je, this_month)
        # Determine kind: prepaid if credit account contains 'prepaid', else accrual
        cred = (a.credit_account or "").lower()
        kind = "prepaid" if "prepaid" in cred else "accrual"

        if derived == "upcoming" and a.service_period == this_month:
            upcoming_this_month_amount += a.amount
        if derived in ("recognized", "overdue") and kind == "accrual":
            outstanding_accrual_amount += a.amount
        if derived in ("upcoming", "pending") and kind == "prepaid":
            prepaid_balance += a.amount

        d = AccruedExpenseRead.model_validate(a).model_dump()
        d["derived_status"] = derived
        d["kind"] = kind
        rows.append(d)

    total_accrued = sum(a.amount for a in accruals if a.status != AccruedExpenseStatus.cleared)
    pending_count = sum(1 for a in accruals if a.status == AccruedExpenseStatus.accrued)
    cleared_this_month = [
        a for a in accruals
        if a.status == AccruedExpenseStatus.cleared and a.service_period == this_month
    ]

    summary = {
        "total_accrued": round(total_accrued, 2),
        "pending_payment_count": pending_count,
        "cleared_this_month": len(cleared_this_month),
        "cleared_this_month_amount": round(sum(a.amount for a in cleared_this_month), 2),
        # New derived totals
        "outstanding_accruals": round(outstanding_accrual_amount, 2),
        "upcoming_this_month_amount": round(upcoming_this_month_amount, 2),
        "prepaid_balance": round(prepaid_balance, 2),
    }

    return {
        "summary": summary,
        "accruals": rows,
    }


# ── Prepaid Schedule Excel export ─────────────────────────────────────────────

@router.get("/export-prepaid")
def export_prepaid_schedule(
    client_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return an xlsx of the prepaid amortization schedule: vendor rows × month columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    _get_client(client_id, current_user, db)
    rows = (
        db.query(AccruedExpense)
        .filter(AccruedExpense.client_id == client_id)
        .order_by(AccruedExpense.vendor_name, AccruedExpense.service_period)
        .all()
    )
    # Only prepaid items
    prepaid_rows = [a for a in rows if "prepaid" in (a.credit_account or "").lower()]

    # Build month set
    months = sorted({a.service_period for a in prepaid_rows})
    # Group by vendor + description (one row per amortization contract)
    by_key: dict[tuple[str, str], dict] = {}
    for a in prepaid_rows:
        key = (a.vendor_name, (a.description or "")[:80])
        if key not in by_key:
            by_key[key] = {
                "vendor": a.vendor_name,
                "description": a.description or "",
                "expense_account": a.debit_account or "",
                "prepaid_account": a.credit_account or "",
                "amounts": {},
                "statuses": {},
            }
        by_key[key]["amounts"][a.service_period] = a.amount
        je = (
            db.query(JournalEntry).filter(JournalEntry.id == a.accrual_je_id).first()
            if a.accrual_je_id else None
        )
        by_key[key]["statuses"][a.service_period] = _derive_status(a, je, _current_month())

    wb = Workbook()
    ws = wb.active
    ws.title = "Prepaid Schedule"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="334155")
    headers = ["Vendor", "Description", "Expense Account", "Prepaid Account"] + months + ["Total"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    STATUS_FILL = {
        "recognized": PatternFill("solid", fgColor="DCFCE7"),  # green
        "pending":    PatternFill("solid", fgColor="FEF9C3"),  # yellow
        "upcoming":   PatternFill("solid", fgColor="E0E7FF"),  # indigo
        "overdue":    PatternFill("solid", fgColor="FEE2E2"),  # red
        "cleared":    PatternFill("solid", fgColor="D1FAE5"),  # darker green
    }
    for row_idx, key in enumerate(sorted(by_key.keys()), start=2):
        row = by_key[key]
        ws.cell(row=row_idx, column=1, value=row["vendor"])
        ws.cell(row=row_idx, column=2, value=row["description"])
        ws.cell(row=row_idx, column=3, value=row["expense_account"])
        ws.cell(row=row_idx, column=4, value=row["prepaid_account"])
        total = 0.0
        for m_idx, m in enumerate(months, start=5):
            amt = row["amounts"].get(m)
            if amt is None:
                continue
            cell = ws.cell(row=row_idx, column=m_idx, value=round(amt, 2))
            cell.number_format = "$#,##0.00"
            st = row["statuses"].get(m)
            if st and st in STATUS_FILL:
                cell.fill = STATUS_FILL[st]
            total += amt
        tcell = ws.cell(row=row_idx, column=len(headers), value=round(total, 2))
        tcell.number_format = "$#,##0.00"
        tcell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    for col_idx in range(5, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12

    # Legend rows under the table
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    for i, (label, fill) in enumerate(STATUS_FILL.items()):
        c = ws.cell(row=legend_row, column=2 + i, value=label.title())
        c.fill = fill
        c.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"prepaid-schedule-{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Create ────────────────────────────────────────────────────────────────────

@router.post("")
def create_accrual(
    client_id: int,
    body: AccruedExpenseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)

    # Determine the accrual date (last day of service period month)
    try:
        sp_dt = datetime.strptime(body.service_period, "%Y-%m")
    except ValueError:
        raise HTTPException(400, "service_period must be YYYY-MM")

    import calendar
    last_day = calendar.monthrange(sp_dt.year, sp_dt.month)[1]
    accrual_date = datetime(sp_dt.year, sp_dt.month, last_day)

    # Create synthetic transaction to hold the accrual JE
    tx = _create_synthetic_transaction(client_id, body.vendor_name, body.amount, accrual_date, db)

    # Create the accrual JE: DR expense / CR accrued expenses
    accrual_je = _create_accrual_je(
        tx_id=tx.id,
        expense_account=body.expense_account,
        accrued_account=body.accrued_account,
        amount=body.amount,
        je_date=accrual_date,
        vendor=body.vendor_name,
        service_period=body.service_period,
        confidence=1.0,
        reasoning="Manually created accrual.",
        db=db,
    )

    ae = AccruedExpense(
        client_id=client_id,
        vendor_name=body.vendor_name,
        description=body.description,
        service_period=body.service_period,
        amount=body.amount,
        source_transaction_id=body.source_transaction_id,
        accrual_je_id=accrual_je.id,
        expected_payment_date=body.expected_payment_date,
        status=AccruedExpenseStatus.accrued,
        ai_confidence=1.0,
        ai_reasoning="Manually created.",
    )
    db.add(ae)

    # If a standing rule was waiting on the user to create this accrual
    # (vendor + service period match, attention flag set), clear the flag
    # — the user just resolved it.
    flagged_rule = (
        db.query(StandingAccrualRule)
        .filter(
            StandingAccrualRule.client_id == client_id,
            StandingAccrualRule.vendor_name == body.vendor_name,
            StandingAccrualRule.attention_needed == True,
            StandingAccrualRule.attention_month == body.service_period,
        )
        .first()
    )
    if flagged_rule:
        flagged_rule.attention_needed = False
        flagged_rule.attention_month  = None
        flagged_rule.attention_reason = None
        flagged_rule.last_generated   = body.service_period

    db.commit()
    db.refresh(ae)
    return AccruedExpenseRead.model_validate(ae)


# ── Update ────────────────────────────────────────────────────────────────────

@router.patch("/{ae_id}")
def update_accrual(
    client_id: int,
    ae_id: int,
    body: AccruedExpenseUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)
    ae = db.query(AccruedExpense).filter(
        AccruedExpense.id == ae_id, AccruedExpense.client_id == client_id
    ).first()
    if not ae:
        raise HTTPException(404, "Accrued expense not found")

    if body.status is not None:
        try:
            ae.status = AccruedExpenseStatus(body.status)
        except ValueError:
            raise HTTPException(400, f"Invalid status: {body.status}")
    if body.expected_payment_date is not None:
        ae.expected_payment_date = body.expected_payment_date
    if body.description is not None:
        ae.description = body.description
    if body.amount is not None:
        ae.amount = body.amount
    if body.vendor_name is not None:
        ae.vendor_name = body.vendor_name
    if body.service_period is not None:
        ae.service_period = body.service_period
    if body.debit_account is not None:
        ae.debit_account = body.debit_account
    if body.credit_account is not None:
        ae.credit_account = body.credit_account

    db.commit()
    db.refresh(ae)
    return AccruedExpenseRead.model_validate(ae)


# ── Delete ────────────────────────────────────────────────────────────────────

@router.delete("/{ae_id}")
def delete_accrual(
    client_id: int,
    ae_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)
    ae = db.query(AccruedExpense).filter(
        AccruedExpense.id == ae_id, AccruedExpense.client_id == client_id
    ).first()
    if not ae:
        raise HTTPException(404, "Accrued expense not found")
    db.delete(ae)
    db.commit()
    return {"ok": True}


# ── Release prepaid accrual to Review Queue ───────────────────────────────────

class RegisterExistingJEBody(BaseModel):
    je_id:           int
    service_period:  Optional[str] = None  # "YYYY-MM" — defaults to JE date's month


@router.post("/register-existing-je")
def register_existing_je_as_accrual(
    client_id: int,
    body: RegisterExistingJEBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Register an already-existing journal entry (e.g. one created by an invoice
    upload before the auto-link logic existed) as an AccruedExpense so it shows
    on the Accruals tab. Does NOT create a new transaction or JE — just inserts
    the tracking row. Idempotent: skipped if an AccruedExpense already points
    at this JE.
    """
    _get_client(client_id, current_user, db)
    je = db.query(JournalEntry).filter(JournalEntry.id == body.je_id).first()
    if not je:
        raise HTTPException(404, "JE not found")

    tx = db.query(Transaction).filter(Transaction.id == je.transaction_id).first()
    if not tx or tx.client_id != client_id:
        raise HTTPException(404, "Transaction for this JE not found in this client")

    existing = db.query(AccruedExpense).filter(AccruedExpense.accrual_je_id == je.id).first()
    if existing:
        return AccruedExpenseRead.model_validate(existing)

    period = body.service_period or (je.je_date or tx.date).strftime("%Y-%m")
    ae = AccruedExpense(
        client_id=client_id,
        vendor_name=tx.counterparty_name or tx.description[:80],
        description=(tx.description or "")[:255],
        service_period=period,
        amount=abs(float(je.amount or 0)),
        source_transaction_id=tx.id,
        accrual_je_id=je.id,
        debit_account=je.debit_account,
        credit_account=je.credit_account,
        status=AccruedExpenseStatus.accrued,
        ai_confidence=1.0,
        ai_reasoning="Linked from existing JE (retroactive registration).",
    )
    db.add(ae)
    db.commit()
    db.refresh(ae)
    return AccruedExpenseRead.model_validate(ae)


@router.post("/{ae_id}/release")
def release_accrual(
    client_id: int,
    ae_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Create a pending transaction + JE for a prepaid amortization accrual,
    making it appear in the Review Queue for the current month.
    """
    _get_client(client_id, current_user, db)
    ae = db.query(AccruedExpense).filter(
        AccruedExpense.id == ae_id, AccruedExpense.client_id == client_id
    ).first()
    if not ae:
        raise HTTPException(404, "Accrued expense not found")
    if ae.accrual_je_id:
        raise HTTPException(400, "Accrual already released to Review Queue")
    if not ae.debit_account or not ae.credit_account:
        raise HTTPException(400, "Accrual missing account info — cannot create JE")

    # Parse the service period date
    try:
        period_dt = datetime.strptime(ae.service_period, "%Y-%m")
    except ValueError:
        period_dt = datetime.utcnow()

    # Last day of the service period month
    import calendar
    last_day = calendar.monthrange(period_dt.year, period_dt.month)[1]
    je_date = period_dt.replace(day=last_day)

    # Create a pending transaction for this month's amortization
    tx = Transaction(
        client_id=client_id,
        date=je_date,
        description=f"{ae.vendor_name} prepaid amortization ({ae.service_period})",
        counterparty_name=ae.vendor_name,
        amount=-abs(ae.amount),
        status=TransactionStatus.pending,
        source="accrual",
    )
    db.add(tx)
    db.flush()

    # Create the JE
    je = JournalEntry(
        transaction_id=tx.id,
        je_number=next_je_number(db),
        debit_account=ae.debit_account,
        credit_account=ae.credit_account,
        amount=ae.amount,
        je_date=je_date,
        memo=f"{ae.vendor_name} amortization ({ae.service_period})"[:80],
        ai_confidence=ae.ai_confidence or 0.9,
        ai_reasoning=ae.ai_reasoning or "",
    )
    db.add(je)
    db.flush()

    ae.accrual_je_id = je.id
    db.commit()

    return {
        "accrual_id": ae.id,
        "transaction_id": tx.id,
        "je_id": je.id,
        "service_period": ae.service_period,
    }


# ── Auto-release (called by scheduler + manual endpoint) ─────────────────────

def _auto_release_for_client(client_id: int, target_month: str, db: Session) -> dict:
    """
    Release all unreleased items for `target_month` (YYYY-MM) into the review queue:
      1. AccruedExpense records (prepaid amortization, manual accruals)
      2. FixedAsset depreciation / amortization
      3. Revenue schedule entries

    Returns a dict with counts of what was created.
    """
    try:
        period_dt = datetime.strptime(target_month, "%Y-%m")
    except ValueError:
        raise ValueError(f"target_month must be YYYY-MM, got: {target_month}")

    last_day_num = calendar.monthrange(period_dt.year, period_dt.month)[1]
    je_date = period_dt.replace(day=last_day_num)

    released_accruals = 0
    released_depreciation = 0
    released_revenue = 0

    # ── 1. Prepaid / manual accruals ──────────────────────────────────────────
    unreleased = db.query(AccruedExpense).filter(
        AccruedExpense.client_id == client_id,
        AccruedExpense.service_period == target_month,
        AccruedExpense.accrual_je_id == None,
        AccruedExpense.debit_account != None,
        AccruedExpense.credit_account != None,
    ).all()

    for ae in unreleased:
        tx = Transaction(
            client_id=client_id,
            date=je_date,
            description=f"{ae.vendor_name} accrual release ({ae.service_period})",
            counterparty_name=ae.vendor_name,
            amount=-abs(ae.amount),
            status=TransactionStatus.pending,
            source="accrual",
        )
        db.add(tx)
        db.flush()

        je = JournalEntry(
            transaction_id=tx.id,
            je_number=next_je_number(db),
            debit_account=ae.debit_account,
            credit_account=ae.credit_account,
            amount=ae.amount,
            je_date=je_date,
            memo=f"{ae.vendor_name} accrual ({ae.service_period})"[:80],
            ai_confidence=ae.ai_confidence or 0.9,
            ai_reasoning=ae.ai_reasoning or "",
        )
        db.add(je)
        db.flush()

        ae.accrual_je_id = je.id
        released_accruals += 1

    # ── 2. Fixed asset depreciation / amortization ───────────────────────────
    from routers.fixed_assets import _compute_schedule, _create_dep_transaction

    assets = db.query(FixedAsset).filter(
        FixedAsset.client_id == client_id,
        FixedAsset.status == "active",
    ).all()

    for asset in assets:
        schedule = _compute_schedule(asset)
        period_entry = next((p for p in schedule if p["period"] == target_month), None)
        if not period_entry:
            continue

        # Skip if depreciation tx already exists for this asset + month
        existing = db.query(Transaction).filter(
            Transaction.fixed_asset_id == asset.id,
            Transaction.source == "depreciation",
        ).all()
        existing_months = {
            f"{t.date.year}-{t.date.month:02d}" for t in existing
        }
        if target_month in existing_months:
            continue

        _create_dep_transaction(asset, period_entry, client_id, db)

        # Mark fully depreciated if this is the last period
        if schedule and schedule[-1]["period"] <= target_month:
            asset.status = "fully_depreciated"

        released_depreciation += 1

    # ── 3. Revenue schedule entries ──────────────────────────────────────────
    pending_entries = db.query(RevenueScheduleEntry).filter(
        RevenueScheduleEntry.client_id == client_id,
        RevenueScheduleEntry.period == target_month,
        RevenueScheduleEntry.je_id == None,
    ).all()

    for entry in pending_entries:
        contract = db.query(RevenueContract).filter(
            RevenueContract.id == entry.contract_id
        ).first()
        if not contract or not contract.revenue_stream_id:
            continue

        stream = db.query(RevenueStream).filter(
            RevenueStream.id == contract.revenue_stream_id
        ).first()
        if not stream:
            continue

        billing_type = (
            BillingType(stream.billing_type)
            if isinstance(stream.billing_type, str)
            else stream.billing_type
        )

        if billing_type in (BillingType.monthly_arrears, BillingType.invoice_completion):
            debit_acct = stream.ar_account
        else:
            debit_acct = stream.deferred_revenue_account
        credit_acct = stream.revenue_account
        memo = f"Revenue Recognition - {contract.customer_name} - {period_dt.strftime('%b %Y')}"

        tx = Transaction(
            client_id=client_id,
            date=je_date,
            description=memo,
            amount=entry.amount,
            status=TransactionStatus.pending,
            source="revenue",
        )
        db.add(tx)
        db.flush()

        je = JournalEntry(
            transaction_id=tx.id,
            je_number=next_je_number(db),
            debit_account=debit_acct,
            credit_account=credit_acct,
            amount=entry.amount,
            je_date=je_date,
            memo=memo[:80],
            ai_confidence=1.0,
            ai_reasoning=f"Auto-released: ASC 606 recognition for {contract.customer_name}, {target_month}.",
        )
        db.add(je)
        db.flush()

        entry.je_id = je.id
        contract.amount_recognized = round(
            (contract.amount_recognized or 0.0) + entry.amount, 2
        )
        released_revenue += 1

    db.commit()
    return {
        "month": target_month,
        "released_accruals": released_accruals,
        "released_depreciation": released_depreciation,
        "released_revenue": released_revenue,
    }


@router.post("/auto-release")
def auto_release(
    client_id: int,
    month: Optional[str] = None,  # "YYYY-MM" — defaults to current month
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Manually trigger auto-release for a given month (defaults to current month).
    Releases prepaid accruals, fixed asset depreciation, and revenue schedule entries
    that haven't been pushed to the review queue yet.
    """
    _get_client(client_id, current_user, db)
    target_month = month or _current_month()
    return _auto_release_for_client(client_id, target_month, db)


# ── AI analyze ────────────────────────────────────────────────────────────────

@router.post("/analyze")
def analyze_payments(
    client_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    AI-scan pending and recent Mercury payments and suggest accrual entries.
    Returns suggestions; does NOT create anything automatically.
    """
    client = _get_client(client_id, current_user, db)

    # Pull pending + last 30 days of sent payments that don't already have accruals
    existing_source_ids = {
        ae.source_transaction_id
        for ae in db.query(AccruedExpense).filter(AccruedExpense.client_id == client_id).all()
        if ae.source_transaction_id
    }

    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=60)
    transactions = (
        db.query(Transaction)
        .filter(
            Transaction.client_id == client_id,
            Transaction.date >= cutoff,
            Transaction.amount < 0,  # outgoing payments only
            ~Transaction.id.in_(existing_source_ids),
        )
        .order_by(Transaction.date.desc())
        .limit(50)
        .all()
    )

    if not transactions:
        return {"suggestions": []}

    import ai_coder
    suggestions = ai_coder.analyze_for_accrual(transactions, client)
    return {"suggestions": [s for s in suggestions if s.get("needs_accrual")]}


@router.post("/from-suggestion")
def create_from_suggestion(
    client_id: int,
    body: AccruedExpenseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Accept an AI suggestion and create the accrual JE + AccruedExpense record."""
    return create_accrual(client_id, body, db, current_user)


# ── Standing rules ────────────────────────────────────────────────────────────

@router.get("/standing-rules")
def list_standing_rules(
    client_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)
    rules = (
        db.query(StandingAccrualRule)
        .filter(StandingAccrualRule.client_id == client_id)
        .order_by(StandingAccrualRule.vendor_name)
        .all()
    )
    return [StandingAccrualRuleRead.model_validate(r) for r in rules]


@router.post("/standing-rules")
def create_standing_rule(
    client_id: int,
    body: StandingAccrualRuleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)
    rule = StandingAccrualRule(
        client_id=client_id,
        vendor_name=body.vendor_name,
        description=body.description,
        expense_account=body.expense_account,
        accrued_account=body.accrued_account,
        amount=body.amount,
        active=True,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return StandingAccrualRuleRead.model_validate(rule)


@router.patch("/standing-rules/{rule_id}")
def update_standing_rule(
    client_id: int,
    rule_id: int,
    body: StandingAccrualRuleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)
    rule = db.query(StandingAccrualRule).filter(
        StandingAccrualRule.id == rule_id, StandingAccrualRule.client_id == client_id
    ).first()
    if not rule:
        raise HTTPException(404, "Standing rule not found")

    for field in ("vendor_name", "description", "expense_account", "accrued_account", "amount", "active", "last_generated"):
        val = getattr(body, field, None)
        if val is not None:
            setattr(rule, field, val)

    db.commit()
    db.refresh(rule)
    return StandingAccrualRuleRead.model_validate(rule)


# ── Accrual Setup (payment processing with split JEs) ─────────────────────────

@router.get("/setup-context")
def get_setup_context(
    client_id: int,
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return open accruals + standing rule context for a given payment transaction."""
    _get_client(client_id, current_user, db)

    tx = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.client_id == client_id,
    ).first()
    if not tx:
        raise HTTPException(404, "Transaction not found")

    vendor = tx.counterparty_name or tx.description or ""

    # Find matching standing rules (case-insensitive vendor match)
    rules = (
        db.query(StandingAccrualRule)
        .filter(
            StandingAccrualRule.client_id == client_id,
            StandingAccrualRule.active == True,
        )
        .all()
    )
    matching_rules = [
        r for r in rules
        if vendor.lower() in r.vendor_name.lower() or r.vendor_name.lower() in vendor.lower()
    ]

    # Find open accruals for any matching vendor names
    vendor_names = {r.vendor_name for r in matching_rules} or {vendor}
    open_accruals = (
        db.query(AccruedExpense)
        .filter(
            AccruedExpense.client_id == client_id,
            AccruedExpense.status.in_([AccruedExpenseStatus.accrued, AccruedExpenseStatus.partially_paid]),
            AccruedExpense.vendor_name.in_(list(vendor_names)),
        )
        .order_by(AccruedExpense.service_period.asc())
        .all()
    )

    return {
        "transaction": {
            "id": tx.id,
            "date": tx.date.isoformat() if tx.date else None,
            "amount": tx.amount,
            "vendor": vendor,
            "mercury_account_name": tx.mercury_account_name,
        },
        "open_accruals": [AccruedExpenseRead.model_validate(a) for a in open_accruals],
        "matching_rules": [StandingAccrualRuleRead.model_validate(r) for r in matching_rules],
    }


@router.post("/setup-payment")
def setup_payment(
    client_id: int,
    body: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Process an accrual payment with split JEs.

    Body:
      transaction_id: int
      clearings: [{accrual_id: int, amount: float}]
      prepaid: optional {
        account: str,          # e.g. "Prepaid Insurance"
        monthly_amount: float,
        start_period: str,     # "YYYY-MM"
        end_period: str,       # "YYYY-MM"
        description: str,
        expense_account: str,  # DR each month, e.g. "Officers' life insurance"
      }
      bank_account: str  # default "Mercury Checking"
    """
    _get_client(client_id, current_user, db)

    transaction_id = body.get("transaction_id")
    clearings = body.get("clearings", [])   # [{accrual_id, amount}]
    prepaid = body.get("prepaid")           # optional prepaid config
    bank_account = body.get("bank_account", "Mercury Checking")

    tx = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.client_id == client_id,
    ).first()
    if not tx:
        raise HTTPException(404, "Transaction not found")

    total_payment = abs(tx.amount)

    # Validate amounts
    clearing_total = sum(c["amount"] for c in clearings)
    prepaid_amount = prepaid["amount"] if prepaid else 0.0
    if abs(clearing_total + prepaid_amount - total_payment) > 0.02:
        raise HTTPException(400, f"Split amounts ({clearing_total + prepaid_amount:.2f}) do not equal transaction amount ({total_payment:.2f})")

    # Delete existing JEs on this transaction
    db.query(JournalEntry).filter(JournalEntry.transaction_id == transaction_id).delete()
    db.flush()

    pay_date = tx.date if tx.date else datetime.utcnow()
    created_jes = []

    # Create a clearing JE for each accrual
    for c in clearings:
        accrual_id = c["accrual_id"]
        amount = float(c["amount"])

        ae = db.query(AccruedExpense).filter(
            AccruedExpense.id == accrual_id,
            AccruedExpense.client_id == client_id,
        ).first()
        if not ae:
            raise HTTPException(404, f"AccruedExpense {accrual_id} not found")

        # Find the accrued_account from the standing rule if available
        accrued_account = "Accrued Expenses"
        if ae.standing_rule_id:
            rule = db.query(StandingAccrualRule).filter(StandingAccrualRule.id == ae.standing_rule_id).first()
            if rule:
                accrued_account = rule.accrued_account

        je = JournalEntry(
            transaction_id=transaction_id,
            je_number=next_je_number(db),
            debit_account=accrued_account,
            credit_account=bank_account,
            amount=amount,
            je_date=pay_date,
            memo=f"Clear accrual: {ae.vendor_name} ({ae.service_period})"[:80],
            ai_confidence=1.0,
            ai_reasoning="Accrual payment setup.",
        )
        db.add(je)
        db.flush()
        created_jes.append(je.id)

        # Update accrual status
        if abs(amount - ae.amount) < 0.01:
            ae.status = AccruedExpenseStatus.cleared
            ae.payment_je_id = je.id
        else:
            ae.status = AccruedExpenseStatus.partially_paid
            ae.amount = round(ae.amount - amount, 2)

    # Create prepaid JE if configured
    amort_transactions: list = []
    if prepaid:
        prepaid_account = prepaid.get("account", "Prepaid Insurance")
        prepaid_amt = float(prepaid.get("amount", prepaid_amount))
        prepaid_desc = prepaid.get("description", f"Prepaid: {tx.counterparty_name or tx.description}")

        je = JournalEntry(
            transaction_id=transaction_id,
            je_number=next_je_number(db),
            debit_account=prepaid_account,
            credit_account=bank_account,
            amount=prepaid_amt,
            je_date=pay_date,
            memo=prepaid_desc[:80],
            ai_confidence=1.0,
            ai_reasoning="Prepaid expense from accrual payment setup.",
        )
        db.add(je)
        db.flush()
        created_jes.append(je.id)

        # Generate monthly amortization entries if start/end period provided
        start_period = prepaid.get("start_period")
        end_period = prepaid.get("end_period")
        expense_account = prepaid.get("expense_account", "Officers' life insurance")
        monthly_amount = float(prepaid.get("monthly_amount", 0))

        if start_period and end_period and monthly_amount:
            try:
                cur = datetime.strptime(start_period, "%Y-%m")
                end = datetime.strptime(end_period, "%Y-%m")
            except ValueError:
                raise HTTPException(400, "start_period and end_period must be YYYY-MM")

            while cur <= end:
                last_day = calendar.monthrange(cur.year, cur.month)[1]
                amort_date = datetime(cur.year, cur.month, last_day)
                period_str = cur.strftime("%Y-%m")

                amort_tx = Transaction(
                    client_id=client_id,
                    date=amort_date,
                    description=f"Amortize prepaid: {tx.counterparty_name or 'prepaid'} ({period_str})",
                    amount=-monthly_amount,
                    status=TransactionStatus.pending,
                    source="accrual",
                )
                db.add(amort_tx)
                db.flush()

                amort_je = JournalEntry(
                    transaction_id=amort_tx.id,
                    je_number=next_je_number(db),
                    debit_account=expense_account,
                    credit_account=prepaid_account,
                    amount=monthly_amount,
                    je_date=amort_date,
                    memo=f"Prepaid amortization: {period_str}"[:80],
                    ai_confidence=1.0,
                    ai_reasoning="Auto-generated prepaid amortization from accrual setup.",
                )
                db.add(amort_je)
                db.flush()
                amort_transactions.append({"period": period_str, "tx_id": amort_tx.id, "je_id": amort_je.id})

                # advance month
                if cur.month == 12:
                    cur = cur.replace(year=cur.year + 1, month=1)
                else:
                    cur = cur.replace(month=cur.month + 1)

    db.commit()

    result: dict = {
        "transaction_id": transaction_id,
        "jes_created": len(created_jes),
        "je_ids": created_jes,
    }
    if prepaid:
        result["amortization_entries"] = len(amort_transactions)
    return result


@router.delete("/standing-rules/{rule_id}")
def delete_standing_rule(
    client_id: int,
    rule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_client(client_id, current_user, db)
    rule = db.query(StandingAccrualRule).filter(
        StandingAccrualRule.id == rule_id, StandingAccrualRule.client_id == client_id
    ).first()
    if not rule:
        raise HTTPException(404, "Standing rule not found")
    db.delete(rule)
    db.commit()
    return {"ok": True}


def _generate_one_month(client_id: int, rule: StandingAccrualRule, target_month: str, db: Session) -> AccruedExpense:
    """Create the synthetic tx + accrual JE + AccruedExpense row for a single
    month of a standing rule. Transaction status is `pending` when the
    accrual_date is on/before today (user should review) and `approved` for
    future months (so they show up in the prepaid schedule but don't clutter
    the Review Queue)."""
    import calendar
    sp_dt = datetime.strptime(target_month, "%Y-%m")
    last_day = calendar.monthrange(sp_dt.year, sp_dt.month)[1]
    accrual_date = datetime(sp_dt.year, sp_dt.month, last_day)
    tx_status = (
        TransactionStatus.pending
        if accrual_date <= datetime.utcnow()
        else TransactionStatus.approved
    )
    tx = _create_synthetic_transaction(
        client_id, rule.vendor_name, rule.amount, accrual_date, db, status=tx_status,
    )
    accrual_je = _create_accrual_je(
        tx_id=tx.id,
        expense_account=rule.expense_account,
        accrued_account=rule.accrued_account,
        amount=rule.amount,
        je_date=accrual_date,
        vendor=rule.vendor_name,
        service_period=target_month,
        confidence=1.0,
        reasoning=f"Standing accrual rule: {rule.description or rule.vendor_name}",
        db=db,
    )
    ae = AccruedExpense(
        client_id=client_id,
        vendor_name=rule.vendor_name,
        description=rule.description,
        service_period=target_month,
        amount=rule.amount,
        accrual_je_id=accrual_je.id,
        status=AccruedExpenseStatus.accrued,
        ai_confidence=1.0,
        ai_reasoning=f"Generated from standing rule #{rule.id}.",
        standing_rule_id=rule.id,
        debit_account=rule.expense_account,
        credit_account=rule.accrued_account,
    )
    db.add(ae)
    return ae


@router.post("/standing-rules/generate")
def generate_from_standing_rules(
    client_id: int,
    month: Optional[str] = None,  # "YYYY-MM", defaults to current month
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate accrual JEs for all active standing rules.

    For an open-ended rule, generates just the target month.
    For a rule with schedule_end_month set (fixed-window prepaid amortization),
    generates every month from last_generated+1 (or the target month, whichever
    is earlier) through schedule_end_month — so the full amortization schedule
    becomes visible immediately.
    """
    _get_client(client_id, current_user, db)
    target_month = month or _current_month()

    try:
        datetime.strptime(target_month, "%Y-%m")
    except ValueError:
        raise HTTPException(400, "month must be YYYY-MM")

    rules = (
        db.query(StandingAccrualRule)
        .filter(
            StandingAccrualRule.client_id == client_id,
            StandingAccrualRule.active == True,
        )
        .all()
    )

    generated = []
    skipped = []

    for rule in rules:
        if rule.amount is None:
            rule.attention_needed = True
            rule.attention_month  = target_month
            rule.attention_reason = (
                "Variable-amount vendor: create the accrual for this "
                "month manually with the actual invoice/estimate amount."
            )
            skipped.append(f"{rule.vendor_name} (no fixed amount — flagged for review)")
            continue

        # Determine which months to generate. For a fixed-window prepaid
        # amortization, generate everything from (last_generated+1 or target,
        # whichever comes first) through schedule_end_month.
        if rule.schedule_end_month:
            window_start = (
                _next_month_str(rule.last_generated)
                if rule.last_generated
                else target_month
            )
            months_to_generate = _months_in_range(window_start, rule.schedule_end_month)
        else:
            months_to_generate = (
                [] if rule.last_generated == target_month else [target_month]
            )

        if not months_to_generate:
            skipped.append(f"{rule.vendor_name} (already complete)")
            continue

        generated_for_rule: list[str] = []
        for m in months_to_generate:
            _generate_one_month(client_id, rule, m, db)
            generated_for_rule.append(m)
            rule.last_generated = m

        rule.attention_needed = False
        rule.attention_month  = None
        rule.attention_reason = None
        generated.append(
            f"{rule.vendor_name} ({', '.join(generated_for_rule)})"
            if len(generated_for_rule) > 1
            else rule.vendor_name
        )

    db.commit()
    return {"generated": generated, "skipped": skipped, "month": target_month}
