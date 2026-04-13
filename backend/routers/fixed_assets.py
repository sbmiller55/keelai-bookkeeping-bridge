import calendar
import io
from datetime import datetime, date
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
import models
import schemas
from models import next_je_number

router = APIRouter(prefix="/clients", tags=["fixed_assets"])

# ── Category keyword → (category, useful_life_months, method) ─────────────────

_TANGIBLE_RULES = [
    (["laptop", "macbook", "macmini", "imac", "computer", "server", "workstation",
      "monitor", "display", "printer", "scanner", "projector", "hardware", "ipad",
      "tablet", "camera"], "Equipment", 60, "straight_line"),
    (["iphone", "android", "phone"], "Equipment", 36, "straight_line"),
    (["furniture", "chair", "desk", "table", "shelving", "cabinet", "sofa", "couch"], "Furniture", 84, "straight_line"),
    (["vehicle", "car", "truck", "van", "automobile"], "Vehicles", 60, "straight_line"),
    (["renovation", "improvement", "remodel", "construction", "build-out", "buildout",
      "leasehold"], "Leasehold Improvements", 120, "straight_line"),
]

_INTANGIBLE_RULES = [
    (["goodwill"], "Goodwill", 0, "straight_line"),                                   # indefinite life
    (["patent"], "Patents", 240, "straight_line"),                                    # 20 years
    (["trademark", "trade mark", "trade-mark"], "Trademarks", 120, "straight_line"), # 10 years
    (["customer list", "customer relationship"], "Customer Lists", 60, "straight_line"),
    (["non-compete", "noncompete", "non compete"], "Non-Compete Agreements", 36, "straight_line"),
    (["software", "saas"], "Capitalized Software", 36, "straight_line"),
    (["license", "licence", "domain"], "Licenses", 36, "straight_line"),
]


def _suggest(description: str) -> dict:
    desc_lower = description.lower()

    for keywords, category, useful_life, method in _INTANGIBLE_RULES:
        if any(kw in desc_lower for kw in keywords):
            return {
                "name": description,
                "category": category,
                "useful_life_months": useful_life,
                "depreciation_method": method,
                "salvage_value": 0.0,
                "asset_type": "intangible",
                "is_indefinite_life": category == "Goodwill",
            }

    for keywords, category, useful_life, method in _TANGIBLE_RULES:
        if any(kw in desc_lower for kw in keywords):
            return {
                "name": description,
                "category": category,
                "useful_life_months": useful_life,
                "depreciation_method": method,
                "salvage_value": 0.0,
                "asset_type": "tangible",
                "is_indefinite_life": False,
            }

    return {
        "name": description,
        "category": "Equipment",
        "useful_life_months": 60,
        "depreciation_method": "straight_line",
        "salvage_value": 0.0,
        "asset_type": "tangible",
        "is_indefinite_life": False,
    }


# ── Depreciation schedule computation ────────────────────────────────────────

def _compute_schedule(asset: models.FixedAsset) -> list[dict]:
    try:
        purchase = datetime.strptime(asset.purchase_date, "%Y-%m-%d").date()
    except Exception:
        return []

    if purchase.month == 12:
        start_year, start_month = purchase.year + 1, 1
    else:
        start_year, start_month = purchase.year, purchase.month + 1

    if getattr(asset, "is_indefinite_life", False) or asset.useful_life_months <= 0:
        return []

    depreciable = asset.purchase_price - asset.salvage_value
    nbv = float(asset.purchase_price)
    accum = 0.0
    schedule = []

    if asset.depreciation_method == "double_declining":
        ddb_rate = 2.0 / asset.useful_life_months
        for i in range(asset.useful_life_months):
            y = start_year + (start_month + i - 1) // 12
            m = (start_month + i - 1) % 12 + 1
            remaining = asset.useful_life_months - i
            sl = (nbv - asset.salvage_value) / remaining if remaining > 0 else 0
            dep = round(min(max(sl, nbv * ddb_rate), nbv - asset.salvage_value), 2)
            if dep <= 0:
                break
            accum = round(accum + dep, 2)
            nbv = round(nbv - dep, 2)
            last_day = calendar.monthrange(y, m)[1]
            schedule.append({
                "period": f"{y}-{m:02d}",
                "date": date(y, m, last_day).isoformat(),
                "depreciation": dep,
                "accumulated_depreciation": accum,
                "net_book_value": nbv,
            })
    else:  # straight_line
        if asset.useful_life_months <= 0:
            return []
        monthly = depreciable / asset.useful_life_months
        for i in range(asset.useful_life_months):
            y = start_year + (start_month + i - 1) // 12
            m = (start_month + i - 1) % 12 + 1
            dep = round(min(monthly, nbv - asset.salvage_value), 2)
            if dep <= 0:
                break
            accum = round(accum + dep, 2)
            nbv = round(nbv - dep, 2)
            last_day = calendar.monthrange(y, m)[1]
            schedule.append({
                "period": f"{y}-{m:02d}",
                "date": date(y, m, last_day).isoformat(),
                "depreciation": dep,
                "accumulated_depreciation": accum,
                "net_book_value": nbv,
            })

    return schedule


def _enrich(asset: models.FixedAsset) -> schemas.FixedAssetRead:
    schedule = _compute_schedule(asset)
    today_period = date.today().strftime("%Y-%m")
    accum_to_date = sum(p["depreciation"] for p in schedule if p["period"] <= today_period)
    nbv = round(asset.purchase_price - accum_to_date, 2)
    monthly_dep = schedule[0]["depreciation"] if schedule else 0.0

    r = schemas.FixedAssetRead.model_validate(asset)
    r.schedule = [schemas.DepreciationPeriod(**p) for p in schedule]
    r.accumulated_depreciation_to_date = round(accum_to_date, 2)
    r.net_book_value = nbv
    r.monthly_depreciation = monthly_dep
    return r


# ── Auth helpers ──────────────────────────────────────────────────────────────

def _get_client_or_403(client_id: int, user: models.User, db: Session) -> models.Client:
    client = db.query(models.Client).filter(
        models.Client.id == client_id,
        models.Client.user_id == user.id,
    ).first()
    if not client:
        raise HTTPException(status_code=403, detail="Client not found")
    return client


def _get_asset_or_404(asset_id: int, client_id: int, db: Session) -> models.FixedAsset:
    asset = db.query(models.FixedAsset).filter(
        models.FixedAsset.id == asset_id,
        models.FixedAsset.client_id == client_id,
    ).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/{client_id}/fixed-assets", response_model=List[schemas.FixedAssetRead])
def list_assets(
    client_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_client_or_403(client_id, current_user, db)
    assets = db.query(models.FixedAsset).filter(
        models.FixedAsset.client_id == client_id,
    ).order_by(models.FixedAsset.purchase_date).all()
    return [_enrich(a) for a in assets]


@router.get("/{client_id}/fixed-assets/export")
def export_schedule(
    client_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download the depreciation schedule as an XLSX file."""
    _get_client_or_403(client_id, current_user, db)
    assets = db.query(models.FixedAsset).filter(
        models.FixedAsset.client_id == client_id,
    ).order_by(models.FixedAsset.purchase_date).all()
    enriched = [_enrich(a) for a in assets]

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    total_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=12)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    right      = Alignment(horizontal="right")

    def money(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=round(val, 2) if val else 0)
        c.number_format = '#,##0.00'
        c.alignment = right
        return c

    def hdr_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        return c

    # ── Sheet 1: Summary (one row per asset, one column per year) ────────────
    ws1 = wb.active
    ws1.title = "Schedule"

    # Collect all calendar years
    year_set: set[int] = set()
    for a in enriched:
        for p in a.schedule:
            year_set.add(int(p.period[:4]))
    years = sorted(year_set)

    # Title rows
    ws1.cell(row=1, column=1, value="Depreciation Schedule").font = title_font
    ws1.cell(row=2, column=1, value="For Financial Reporting Only.")
    ws1.merge_cells(start_row=4, start_column=7, end_row=4, end_column=max(7, 6 + len(years)))
    ws1.cell(row=4, column=7, value="Depreciation for Year …").font = Font(italic=True, size=9)
    ws1.cell(row=4, column=7).alignment = center

    # Header row
    headers = ["Asset", "Cost", "Year\nAcquired", "Salvage\nValue", "Life\n(Yrs)", "Method"] + [str(y) for y in years]
    for col, h in enumerate(headers, 1):
        c = hdr_cell(ws1, 5, col, h)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws1.row_dimensions[5].height = 30

    year_totals: dict[int, float] = {y: 0.0 for y in years}
    sub_font = Font(italic=True, size=9, color="888888")
    row = 6
    for asset in enriched:
        # Build annual buckets for dep, accum dep (end of year), nbv (end of year)
        annual_dep:   dict[int, float] = {y: 0.0 for y in years}
        annual_accum: dict[int, float] = {y: 0.0 for y in years}
        annual_nbv:   dict[int, float] = {y: asset.purchase_price for y in years}
        for p in asset.schedule:
            y = int(p.period[:4])
            if y in annual_dep:
                annual_dep[y] = round(annual_dep[y] + p.depreciation, 2)
                annual_accum[y] = p.accumulated_depreciation
                annual_nbv[y]   = p.net_book_value

        method = "SL" if asset.depreciation_method == "straight_line" else "DDB"
        life = "Indefinite" if asset.is_indefinite_life else (
            f"{asset.useful_life_months / 12:.1f}" if asset.useful_life_months else "—"
        )

        # Row 1: Depreciation
        ws1.cell(row=row, column=1, value=asset.name)
        money(ws1, row, 2, asset.purchase_price)
        ws1.cell(row=row, column=3, value=int(asset.purchase_date[:4])).alignment = center
        if asset.is_indefinite_life:
            ws1.cell(row=row, column=4, value="—").alignment = center
        else:
            money(ws1, row, 4, asset.salvage_value)
        ws1.cell(row=row, column=5, value=life).alignment = center
        ws1.cell(row=row, column=6, value=method).alignment = center
        for ci, y in enumerate(years, 7):
            dep = annual_dep[y]
            year_totals[y] = round(year_totals[y] + dep, 2)
            if dep:
                money(ws1, row, ci, dep)
            else:
                ws1.cell(row=row, column=ci, value=" - ").alignment = center

        # Row 2: Accumulated Depreciation
        ws1.cell(row=row+1, column=6, value="Accum. Dep.").font = sub_font
        for ci, y in enumerate(years, 7):
            v = annual_accum[y]
            c = money(ws1, row+1, ci, v) if v else ws1.cell(row=row+1, column=ci, value=" - ")
            c.font = sub_font

        # Row 3: Net Book Value
        ws1.cell(row=row+2, column=6, value="Net Book Value").font = sub_font
        for ci, y in enumerate(years, 7):
            c = money(ws1, row+2, ci, annual_nbv[y])
            c.font = sub_font

        row += 4  # dep + accum + nbv + blank

    # Total row (depreciation only)
    ws1.cell(row=row, column=6, value="Total:").font = total_font
    for ci, y in enumerate(years, 7):
        c = money(ws1, row, ci, year_totals[y])
        c.font = total_font
        c.border = Border(top=thin)

    # Column widths
    col_widths = [34, 13, 10, 11, 9, 8] + [13] * len(years)
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Monthly Detail (pivot — one row per asset, months as columns) ──
    ws2 = wb.create_sheet("Monthly Detail")
    ws2.cell(row=1, column=1, value="Depreciation Schedule — Monthly Detail").font = title_font
    ws2.cell(row=2, column=1, value="For Financial Reporting Only.")

    # Collect all months across all assets
    month_set: set[str] = set()
    for a in enriched:
        for p in a.schedule:
            month_set.add(p.period)
    months = sorted(month_set)

    ws2.merge_cells(start_row=4, start_column=7, end_row=4, end_column=max(7, 6 + len(months)))
    ws2.cell(row=4, column=7, value="Depreciation for Month …").font = Font(italic=True, size=9)
    ws2.cell(row=4, column=7).alignment = center

    # Header row
    m_hdrs = ["Asset", "Cost", "Year\nAcquired", "Salvage\nValue", "Life\n(Yrs)", "Method"] + months
    for col, h in enumerate(m_hdrs, 1):
        c = hdr_cell(ws2, 5, col, h)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws2.row_dimensions[5].height = 30

    month_totals: dict[str, float] = {m: 0.0 for m in months}
    row2 = 6
    for asset in enriched:
        monthly_dep:   dict[str, float] = {m: 0.0 for m in months}
        monthly_accum: dict[str, float] = {m: 0.0 for m in months}
        monthly_nbv:   dict[str, float] = {m: asset.purchase_price for m in months}
        for p in asset.schedule:
            if p.period in monthly_dep:
                monthly_dep[p.period]   = p.depreciation
                monthly_accum[p.period] = p.accumulated_depreciation
                monthly_nbv[p.period]   = p.net_book_value

        method = "SL" if asset.depreciation_method == "straight_line" else "DDB"
        life = "Indefinite" if asset.is_indefinite_life else (
            f"{asset.useful_life_months / 12:.1f}" if asset.useful_life_months else "—"
        )

        # Row 1: Depreciation
        ws2.cell(row=row2, column=1, value=asset.name)
        money(ws2, row2, 2, asset.purchase_price)
        ws2.cell(row=row2, column=3, value=int(asset.purchase_date[:4])).alignment = center
        if asset.is_indefinite_life:
            ws2.cell(row=row2, column=4, value="—").alignment = center
        else:
            money(ws2, row2, 4, asset.salvage_value)
        ws2.cell(row=row2, column=5, value=life).alignment = center
        ws2.cell(row=row2, column=6, value=method).alignment = center
        for ci, m in enumerate(months, 7):
            dep = monthly_dep[m]
            month_totals[m] = round(month_totals[m] + dep, 2)
            if dep:
                money(ws2, row2, ci, dep)
            else:
                ws2.cell(row=row2, column=ci, value=" - ").alignment = center

        # Row 2: Accumulated Depreciation
        ws2.cell(row=row2+1, column=6, value="Accum. Dep.").font = sub_font
        for ci, m in enumerate(months, 7):
            v = monthly_accum[m]
            c = money(ws2, row2+1, ci, v) if v else ws2.cell(row=row2+1, column=ci, value=" - ")
            c.font = sub_font

        # Row 3: Net Book Value
        ws2.cell(row=row2+2, column=6, value="Net Book Value").font = sub_font
        for ci, m in enumerate(months, 7):
            c = money(ws2, row2+2, ci, monthly_nbv[m])
            c.font = sub_font

        row2 += 4

    # Total row
    ws2.cell(row=row2, column=6, value="Total:").font = total_font
    for ci, m in enumerate(months, 7):
        c = money(ws2, row2, ci, month_totals[m])
        c.font = total_font
        c.border = Border(top=thin)

    # Column widths
    col_widths2 = [34, 13, 10, 11, 9, 8] + [10] * len(months)
    for ci, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    month = date.today().strftime("%Y-%m")
    filename = f"depreciation-schedule-{month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{client_id}/fixed-assets/suggest")
def suggest_asset(
    client_id: int,
    transaction_id: Optional[int] = Query(None),
    name: Optional[str] = Query(None),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_client_or_403(client_id, current_user, db)
    if transaction_id is not None:
        tx = db.query(models.Transaction).filter(
            models.Transaction.id == transaction_id,
            models.Transaction.client_id == client_id,
        ).first()
        if not tx:
            raise HTTPException(status_code=404, detail="Transaction not found")
        suggestion = _suggest(tx.description)
        suggestion["purchase_date"] = tx.date.strftime("%Y-%m-%d")
        suggestion["purchase_price"] = abs(tx.amount)
        return suggestion
    elif name:
        return _suggest(name)
    else:
        raise HTTPException(status_code=422, detail="Provide transaction_id or name")


@router.post("/{client_id}/fixed-assets", response_model=schemas.FixedAssetRead, status_code=201)
def create_asset(
    client_id: int,
    payload: schemas.FixedAssetCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_client_or_403(client_id, current_user, db)

    # Enforce GAAP rules for intangibles
    dep_method = payload.depreciation_method
    salvage = payload.salvage_value
    if payload.asset_type == "intangible":
        dep_method = "straight_line"
        salvage = 0.0

    asset = models.FixedAsset(
        client_id=client_id,
        transaction_id=payload.transaction_id,
        name=payload.name,
        category=payload.category,
        purchase_date=payload.purchase_date,
        purchase_price=payload.purchase_price,
        salvage_value=salvage,
        useful_life_months=payload.useful_life_months,
        depreciation_method=dep_method,
        asset_type=payload.asset_type,
        is_indefinite_life=payload.is_indefinite_life,
        qbo_asset_account=payload.qbo_asset_account,
        qbo_accum_dep_account=payload.qbo_accum_dep_account,
        qbo_dep_expense_account=payload.qbo_dep_expense_account,
        notes=payload.notes,
    )
    db.add(asset)
    db.flush()  # get asset.id

    # Recode the source transaction JE if provided
    if payload.je_id and payload.qbo_asset_account:
        je = db.query(models.JournalEntry).filter(
            models.JournalEntry.id == payload.je_id,
        ).first()
        if je:
            je.debit_account = payload.qbo_asset_account
            je.ai_reasoning = f"Recoded as fixed asset: {payload.name}"

    # Mark the source transaction as approved (not pending) so it leaves the review queue
    if payload.transaction_id:
        tx = db.query(models.Transaction).filter(
            models.Transaction.id == payload.transaction_id,
            models.Transaction.client_id == client_id,
        ).first()
        if tx:
            tx.status = models.TransactionStatus.approved
            tx.fixed_asset_id = asset.id

    db.commit()
    db.refresh(asset)

    # Generate historical depreciation/amortization JEs (skip indefinite-life assets like Goodwill)
    if not payload.is_indefinite_life:
        _generate_historical_dep_jes(asset, client_id, db)

    return _enrich(asset)


def _generate_historical_dep_jes(asset: models.FixedAsset, client_id: int, db: Session):
    """Generate depreciation JEs for all past periods, marked as already exported (assumed in QBO)."""
    schedule = _compute_schedule(asset)
    today_period = date.today().strftime("%Y-%m")

    existing_months = [
        f"{t.date.year}-{t.date.month:02d}"
        for t in db.query(models.Transaction).filter(
            models.Transaction.fixed_asset_id == asset.id,
            models.Transaction.source == "depreciation",
        ).all()
    ]

    for period in schedule:
        if period["period"] >= today_period:
            break  # don't auto-generate current or future months
        if period["period"] in existing_months:
            continue
        _create_dep_transaction(asset, period, client_id, db, status=models.TransactionStatus.exported)

    db.commit()


def _create_dep_transaction(
    asset: models.FixedAsset, period: dict, client_id: int, db: Session,
    status: models.TransactionStatus = models.TransactionStatus.pending,
):
    y, m = int(period["period"].split("-")[0]), int(period["period"].split("-")[1])
    last_day = calendar.monthrange(y, m)[1]
    tx_date = datetime(y, m, last_day)
    month_label = tx_date.strftime("%B %Y")
    term = "Amortization" if getattr(asset, "asset_type", "tangible") == "intangible" else "Depreciation"
    memo = f"{term} - {asset.name} - {month_label}"

    tx = models.Transaction(
        client_id=client_id,
        description=memo,
        amount=period["depreciation"],
        date=tx_date,
        status=status,
        source="depreciation",
        fixed_asset_id=asset.id,
        imported_at=datetime.utcnow(),
    )
    db.add(tx)
    db.flush()

    je = models.JournalEntry(
        transaction_id=tx.id,
        je_number=next_je_number(db),
        debit_account=asset.qbo_dep_expense_account or f"{term} Expense",
        credit_account=asset.qbo_accum_dep_account or f"Accumulated {term}",
        amount=period["depreciation"],
        je_date=tx_date,
        memo=memo,
    )
    db.add(je)


@router.get("/{client_id}/fixed-assets/{asset_id}", response_model=schemas.FixedAssetRead)
def get_asset(
    client_id: int,
    asset_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_client_or_403(client_id, current_user, db)
    asset = _get_asset_or_404(asset_id, client_id, db)
    return _enrich(asset)


@router.put("/{client_id}/fixed-assets/{asset_id}", response_model=schemas.FixedAssetRead)
def update_asset(
    client_id: int,
    asset_id: int,
    payload: schemas.FixedAssetUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_client_or_403(client_id, current_user, db)
    asset = _get_asset_or_404(asset_id, client_id, db)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(asset, field, value)
    db.commit()
    db.refresh(asset)
    return _enrich(asset)


@router.post("/{client_id}/fixed-assets/{asset_id}/dispose", response_model=schemas.FixedAssetRead)
def dispose_asset(
    client_id: int,
    asset_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_client_or_403(client_id, current_user, db)
    asset = _get_asset_or_404(asset_id, client_id, db)
    asset.status = "disposed"
    db.commit()
    db.refresh(asset)
    return _enrich(asset)


@router.post("/{client_id}/fixed-assets/generate-depreciation")
def generate_depreciation(
    client_id: int,
    month: str = Query(...),  # "2026-04"
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _get_client_or_403(client_id, current_user, db)
    assets = db.query(models.FixedAsset).filter(
        models.FixedAsset.client_id == client_id,
        models.FixedAsset.status == "active",
    ).all()

    created = 0
    skipped = 0

    for asset in assets:
        schedule = _compute_schedule(asset)
        period = next((p for p in schedule if p["period"] == month), None)
        if not period:
            skipped += 1
            continue

        # Check for duplicate
        existing_months = [
            f"{t.date.year}-{t.date.month:02d}"
            for t in db.query(models.Transaction).filter(
                models.Transaction.fixed_asset_id == asset.id,
                models.Transaction.source == "depreciation",
            ).all()
        ]
        if month in existing_months:
            skipped += 1
            continue

        _create_dep_transaction(asset, period, client_id, db)

        # Update status if fully depreciated
        if schedule and schedule[-1]["period"] <= month:
            asset.status = "fully_depreciated"

        created += 1

    db.commit()
    return {"created": created, "skipped": skipped, "month": month}
